"""
DiBono ERP - Reports & Analytics Screen
Comprehensive business intelligence with charts and exports
"""

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
                              QTableWidget, QTableWidgetItem, QFrame, QTabWidget,
                              QDateEdit, QComboBox, QMessageBox, QFileDialog, QScrollArea)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtCharts import QChart, QChartView, QBarSet, QBarSeries, QBarCategoryAxis, QValueAxis, QLineSeries, QPieSeries
from PyQt6.QtGui import QPainter, QColor
from models.database import (get_session, Order, OrderItem, MenuItem, User, Session)
from utils.helpers import CurrencyFormatter, DateRangeCalculator
from sqlalchemy import func, extract
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, LineChart, PieChart, Reference


class ReportsManagement(QWidget):
    """Comprehensive reports and analytics"""
    
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.load_money_report()
        
    def init_ui(self):
        """Initialize reports UI"""
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.Shape.NoFrame)
        main_layout.addWidget(scroll_area)

        content_widget = QWidget()
        scroll_area.setWidget(content_widget)
        
        layout = QVBoxLayout(content_widget)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        # Title
        title = QLabel("📊 التقارير والتحليلات")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        layout.addWidget(title)
        
        # Tabs
        self.tabs = QTabWidget()
        
        # Money tab
        money_widget = self.create_money_tab()
        self.tabs.addTab(money_widget, "💰 تحليل الإيرادات")
        
        # Popular items tab
        popular_widget = self.create_popular_items_tab()
        self.tabs.addTab(popular_widget, "🍤 الأصناف الأكثر مبيعًا")
        
        # Peak hours tab
        peak_widget = self.create_peak_hours_tab()
        self.tabs.addTab(peak_widget, "⏰ ساعات الذروة")
        
        # Employee performance tab
        employee_widget = self.create_employee_performance_tab()
        self.tabs.addTab(employee_widget, "👤 أداء الموظفين")
        
        layout.addWidget(self.tabs)
        
        self.tabs.currentChanged.connect(self.on_tab_changed)
        
    def create_money_tab(self):
        """Create money analysis tab"""
        widget = QWidget()
        layout = QVBoxLayout()
        widget.setLayout(layout)
        
        # Date range selector
        controls = QHBoxLayout()
        
        controls.addWidget(QLabel("من:"))
        self.money_from_date = QDateEdit()
        self.money_from_date.setCalendarPopup(True)
        self.money_from_date.setDate(QDate.currentDate().addDays(-30))
        controls.addWidget(self.money_from_date)
        
        controls.addWidget(QLabel("إلى:"))
        self.money_to_date = QDateEdit()
        self.money_to_date.setCalendarPopup(True)
        self.money_to_date.setDate(QDate.currentDate())
        controls.addWidget(self.money_to_date)
        
        refresh_btn = QPushButton("🔄 تحديث")
        refresh_btn.clicked.connect(self.load_money_report)
        controls.addWidget(refresh_btn)
        
        export_btn = QPushButton("📄 تصدير إكسل")
        export_btn.clicked.connect(self.export_money_excel)
        controls.addWidget(export_btn)
        
        controls.addStretch()
        layout.addLayout(controls)
        
        # Stats cards
        stats_frame = QFrame()
        stats_frame.setObjectName("card")
        stats_layout = QHBoxLayout()
        stats_frame.setLayout(stats_layout)
        
        self.money_total_orders = QLabel("الطلبات: 0")
        self.money_total_orders.setStyleSheet("font-size: 12pt; font-weight: bold;")
        stats_layout.addWidget(self.money_total_orders)
        
        self.money_total_revenue = QLabel("الإيراد: 0.00 د.ل")
        self.money_total_revenue.setStyleSheet("font-size: 14pt; font-weight: bold; color: #51CF66;")
        stats_layout.addWidget(self.money_total_revenue)
        
        self.money_avg_order = QLabel("متوسط الطلب: 0.00 د.ل")
        self.money_avg_order.setStyleSheet("font-size: 12pt; font-weight: bold;")
        stats_layout.addWidget(self.money_avg_order)
        
        stats_layout.addStretch()
        layout.addWidget(stats_frame)
        
        # Chart
        self.money_chart = QChart()
        self.money_chart.setTitle("اتجاه الإيرادات اليومية")
        self.money_chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        
        chart_view = QChartView(self.money_chart)
        chart_view.setRenderHint(QPainter.RenderHint.Antialiasing)
        chart_view.setMinimumHeight(300)
        layout.addWidget(chart_view)
        
        # Daily breakdown table
        self.money_table = QTableWidget()
        self.money_table.setColumnCount(5)
        self.money_table.setHorizontalHeaderLabels([
            "التاريخ", "الطلبات", "مبيعات نقدية", "مبيعات بطاقات", "الإجمالي"
        ])
        self.money_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.money_table)
        
        return widget
        
    def create_popular_items_tab(self):
        """Create popular items tab"""
        widget = QWidget()
        layout = QVBoxLayout()
        widget.setLayout(layout)
        
        # Controls
        controls = QHBoxLayout()
        
        controls.addWidget(QLabel("الفترة:"))
        self.popular_period = QComboBox()
        self.popular_period.addItem("اليوم", "today")
        self.popular_period.addItem("هذا الأسبوع", "week")
        self.popular_period.addItem("هذا الشهر", "month")
        self.popular_period.addItem("منذ البداية", "all")
        self.popular_period.currentIndexChanged.connect(self.load_popular_items)
        controls.addWidget(self.popular_period)
        
        refresh_btn = QPushButton("🔄 تحديث")
        refresh_btn.clicked.connect(self.load_popular_items)
        controls.addWidget(refresh_btn)
        
        export_btn = QPushButton("📄 تصدير إكسل")
        export_btn.clicked.connect(self.export_popular_excel)
        controls.addWidget(export_btn)
        
        controls.addStretch()
        layout.addLayout(controls)
        
        # Chart
        self.popular_chart = QChart()
        self.popular_chart.setTitle("أكثر 10 أصناف مبيعًا")
        self.popular_chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        
        chart_view = QChartView(self.popular_chart)
        chart_view.setRenderHint(QPainter.RenderHint.Antialiasing)
        chart_view.setMinimumHeight(300)
        layout.addWidget(chart_view)
        
        # Table
        self.popular_table = QTableWidget()
        self.popular_table.setColumnCount(4)
        self.popular_table.setHorizontalHeaderLabels([
            "الترتيب", "الصنف", "الكمية المباعة", "الإيراد"
        ])
        self.popular_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.popular_table)
        
        return widget
        
    def create_peak_hours_tab(self):
        """Create peak hours tab"""
        widget = QWidget()
        layout = QVBoxLayout()
        widget.setLayout(layout)
        
        # Controls
        controls = QHBoxLayout()
        
        controls.addWidget(QLabel("اليوم:"))
        self.peak_day = QComboBox()
        self.peak_day.addItem("كل الأيام", "All Days")
        self.peak_day.addItem("الاثنين", "Monday")
        self.peak_day.addItem("الثلاثاء", "Tuesday")
        self.peak_day.addItem("الأربعاء", "Wednesday")
        self.peak_day.addItem("الخميس", "Thursday")
        self.peak_day.addItem("الجمعة", "Friday")
        self.peak_day.addItem("السبت", "Saturday")
        self.peak_day.addItem("الأحد", "Sunday")
        self.peak_day.currentIndexChanged.connect(self.load_peak_hours)
        controls.addWidget(self.peak_day)
        
        refresh_btn = QPushButton("🔄 تحديث")
        refresh_btn.clicked.connect(self.load_peak_hours)
        controls.addWidget(refresh_btn)
        
        export_btn = QPushButton("📄 تصدير إكسل")
        export_btn.clicked.connect(self.export_peak_excel)
        controls.addWidget(export_btn)
        
        controls.addStretch()
        layout.addLayout(controls)
        
        # Chart
        self.peak_chart = QChart()
        self.peak_chart.setTitle("الطلبات حسب الساعة")
        self.peak_chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        
        chart_view = QChartView(self.peak_chart)
        chart_view.setRenderHint(QPainter.RenderHint.Antialiasing)
        chart_view.setMinimumHeight(300)
        layout.addWidget(chart_view)
        
        # Table
        self.peak_table = QTableWidget()
        self.peak_table.setColumnCount(3)
        self.peak_table.setHorizontalHeaderLabels([
            "الساعة", "الطلبات", "الإيراد"
        ])
        self.peak_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.peak_table)
        
        return widget
        
    def create_employee_performance_tab(self):
        """Create employee performance tab"""
        widget = QWidget()
        layout = QVBoxLayout()
        widget.setLayout(layout)
        
        # Controls
        controls = QHBoxLayout()
        
        controls.addWidget(QLabel("الفترة:"))
        self.employee_period = QComboBox()
        self.employee_period.addItem("هذا الأسبوع", "week")
        self.employee_period.addItem("هذا الشهر", "month")
        self.employee_period.addItem("منذ البداية", "all")
        self.employee_period.currentIndexChanged.connect(self.load_employee_performance)
        controls.addWidget(self.employee_period)
        
        refresh_btn = QPushButton("🔄 تحديث")
        refresh_btn.clicked.connect(self.load_employee_performance)
        controls.addWidget(refresh_btn)
        
        export_btn = QPushButton("📄 تصدير إكسل")
        export_btn.clicked.connect(self.export_employee_excel)
        controls.addWidget(export_btn)
        
        controls.addStretch()
        layout.addLayout(controls)
        
        # Chart
        self.employee_chart = QChart()
        self.employee_chart.setTitle("مقارنة مبيعات الموظفين")
        self.employee_chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        
        chart_view = QChartView(self.employee_chart)
        chart_view.setRenderHint(QPainter.RenderHint.Antialiasing)
        chart_view.setMinimumHeight(300)
        layout.addWidget(chart_view)
        
        # Table
        self.employee_table = QTableWidget()
        self.employee_table.setColumnCount(5)
        self.employee_table.setHorizontalHeaderLabels([
            "الموظف", "الجلسات", "الطلبات", "إجمالي المبيعات", "متوسط الطلب"
        ])
        self.employee_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.employee_table)
        
        return widget
        
    def on_tab_changed(self, index):
        """Load data when tab changes"""
        if index == 1:  # Popular items
            self.load_popular_items()
        elif index == 2:  # Peak hours
            self.load_peak_hours()
        elif index == 3:  # Employee performance
            self.load_employee_performance()
            
    def load_money_report(self):
        """Load money analysis data"""
        db = get_session()
        try:
            from_date = self.money_from_date.date().toPyDate()
            to_date = self.money_to_date.date().toPyDate()
            to_datetime = datetime.combine(to_date, datetime.max.time())
            
            # Get orders in range
            orders = db.query(Order).filter(
                Order.timestamp >= from_date,
                Order.timestamp <= to_datetime,
                Order.status == 'completed'
            ).order_by(Order.timestamp).all()
            
            # Calculate stats
            total_orders = len(orders)
            total_revenue = sum(o.total for o in orders)
            avg_order = total_revenue / total_orders if total_orders > 0 else 0
            
            self.money_total_orders.setText(f"الطلبات: {total_orders}")
            self.money_total_revenue.setText(f"الإيراد: {CurrencyFormatter.format_lyd(total_revenue)}")
            self.money_avg_order.setText(f"متوسط الطلب: {CurrencyFormatter.format_lyd(avg_order)}")
            
            # Group by date
            daily_data = {}
            for order in orders:
                date_str = order.timestamp.strftime("%Y-%m-%d")
                if date_str not in daily_data:
                    daily_data[date_str] = {'orders': 0, 'cash': 0, 'card': 0, 'total': 0}
                
                daily_data[date_str]['orders'] += 1
                daily_data[date_str]['total'] += order.total
                
                if order.payment_method == 'cash':
                    daily_data[date_str]['cash'] += order.total
                elif order.payment_method == 'card':
                    daily_data[date_str]['card'] += order.total
                elif order.payment_method == 'split':
                    daily_data[date_str]['cash'] += order.cash_amount
                    daily_data[date_str]['card'] += order.card_amount
            
            # Update table
            dates = sorted(daily_data.keys())
            self.money_table.setRowCount(len(dates))
            
            for idx, date in enumerate(dates):
                data = daily_data[date]
                self.money_table.setItem(idx, 0, QTableWidgetItem(date))
                self.money_table.setItem(idx, 1, QTableWidgetItem(str(data['orders'])))
                self.money_table.setItem(idx, 2, QTableWidgetItem(CurrencyFormatter.format_lyd(data['cash'])))
                self.money_table.setItem(idx, 3, QTableWidgetItem(CurrencyFormatter.format_lyd(data['card'])))
                self.money_table.setItem(idx, 4, QTableWidgetItem(CurrencyFormatter.format_lyd(data['total'])))
            
            # Update chart - clear all series and axes first
            self.money_chart.removeAllSeries()
            for axis in self.money_chart.axes():
                self.money_chart.removeAxis(axis)
            
            series = QLineSeries()
            series.setName("الإيراد")
            
            for idx, date in enumerate(dates):
                series.append(idx, CurrencyFormatter.fils_to_lyd(daily_data[date]['total']))
            
            self.money_chart.addSeries(series)
            
            # Axes
            axis_x = QBarCategoryAxis()
            axis_x.append([d[-5:] for d in dates])  # Show MM-DD
            self.money_chart.addAxis(axis_x, Qt.AlignmentFlag.AlignBottom)
            series.attachAxis(axis_x)
            
            axis_y = QValueAxis()
            axis_y.setTitleText("الإيراد (د.ل)")
            self.money_chart.addAxis(axis_y, Qt.AlignmentFlag.AlignLeft)
            series.attachAxis(axis_y)
            
        finally:
            db.close()
            
    def load_popular_items(self):
        """Load popular items data"""
        db = get_session()
        try:
            # Get date range based on period
            period = self.popular_period.currentData()
            if period == "today":
                start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            elif period == "week":
                start_date = DateRangeCalculator.get_week_start()
            elif period == "month":
                start_date = DateRangeCalculator.get_month_start()
            else:
                start_date = datetime(2000, 1, 1)
            
            # Query popular items
            results = db.query(
                MenuItem.name_en,
                MenuItem.name_ar,
                func.sum(OrderItem.quantity).label('total_qty'),
                func.sum(OrderItem.line_total).label('total_revenue')
            ).join(
                OrderItem, OrderItem.menu_item_id == MenuItem.id
            ).join(
                Order, Order.id == OrderItem.order_id
            ).filter(
                Order.timestamp >= start_date,
                Order.status == 'completed'
            ).group_by(
                MenuItem.id
            ).order_by(
                func.sum(OrderItem.quantity).desc()
            ).limit(10).all()
            
            # Update table
            self.popular_table.setRowCount(len(results))
            
            items = []
            quantities = []
            
            for idx, row in enumerate(results):
                self.popular_table.setItem(idx, 0, QTableWidgetItem(f"#{idx + 1}"))
                self.popular_table.setItem(idx, 1, QTableWidgetItem(f"{row.name_ar} / {row.name_en}"))
                
                qty = int(row.total_qty)
                self.popular_table.setItem(idx, 2, QTableWidgetItem(str(qty)))
                self.popular_table.setItem(idx, 3, QTableWidgetItem(CurrencyFormatter.format_lyd(row.total_revenue)))
                
                items.append(row.name_ar[:15])
                quantities.append(qty)
            
            # Update chart - clear all series and axes first
            self.popular_chart.removeAllSeries()
            for axis in self.popular_chart.axes():
                self.popular_chart.removeAxis(axis)
            
            bar_set = QBarSet("الكمية المباعة")
            for qty in quantities:
                bar_set.append(qty)
            
            series = QBarSeries()
            series.append(bar_set)
            
            self.popular_chart.addSeries(series)
            
            axis_x = QBarCategoryAxis()
            axis_x.append(items)
            self.popular_chart.addAxis(axis_x, Qt.AlignmentFlag.AlignBottom)
            series.attachAxis(axis_x)
            
            axis_y = QValueAxis()
            axis_y.setTitleText("الكمية")
            self.popular_chart.addAxis(axis_y, Qt.AlignmentFlag.AlignLeft)
            series.attachAxis(axis_y)
            
        finally:
            db.close()
            
    def load_peak_hours(self):
        """Load peak hours data"""
        db = get_session()
        try:
            # Filter by day if selected
            day_filter = self.peak_day.currentData()
            
            query = db.query(
                extract('hour', Order.timestamp).label('hour'),
                func.count(Order.id).label('order_count'),
                func.sum(Order.total).label('total_revenue')
            ).filter(
                Order.status == 'completed'
            )
            
            if day_filter != "All Days":
                # Map day name to number (0=Monday, 6=Sunday)
                days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
                day_num = days.index(day_filter)
                query = query.filter(extract('dow', Order.timestamp) == (day_num + 1) % 7)
            
            results = query.group_by(
                extract('hour', Order.timestamp)
            ).order_by('hour').all()
            
            # Update table
            self.peak_table.setRowCount(len(results))
            
            hours = []
            counts = []
            
            for idx, row in enumerate(results):
                hour = int(row.hour)
                hour_str = f"{hour:02d}:00 - {hour:02d}:59"
                
                self.peak_table.setItem(idx, 0, QTableWidgetItem(hour_str))
                self.peak_table.setItem(idx, 1, QTableWidgetItem(str(row.order_count)))
                self.peak_table.setItem(idx, 2, QTableWidgetItem(CurrencyFormatter.format_lyd(row.total_revenue)))
                
                hours.append(f"{hour:02d}:00")
                counts.append(int(row.order_count))
            
            # Update chart - clear all series and axes first
            self.peak_chart.removeAllSeries()
            for axis in self.peak_chart.axes():
                self.peak_chart.removeAxis(axis)
            
            bar_set = QBarSet("الطلبات")
            for count in counts:
                bar_set.append(count)
            
            series = QBarSeries()
            series.append(bar_set)
            
            self.peak_chart.addSeries(series)
            
            axis_x = QBarCategoryAxis()
            axis_x.append(hours)
            self.peak_chart.addAxis(axis_x, Qt.AlignmentFlag.AlignBottom)
            series.attachAxis(axis_x)
            
            axis_y = QValueAxis()
            axis_y.setTitleText("عدد الطلبات")
            self.peak_chart.addAxis(axis_y, Qt.AlignmentFlag.AlignLeft)
            series.attachAxis(axis_y)
            
        finally:
            db.close()
            
    def load_employee_performance(self):
        """Load employee performance data"""
        db = get_session()
        try:
            # Get date range
            period = self.employee_period.currentData()
            if period == "week":
                start_date = DateRangeCalculator.get_week_start()
            elif period == "month":
                start_date = DateRangeCalculator.get_month_start()
            else:
                start_date = datetime(2000, 1, 1)
            
            # Query sessions and calculate totals
            results = db.query(
                User.username,
                User.full_name,
                func.count(Session.id).label('session_count'),
                func.sum(Session.orders_count).label('total_orders'),
                func.sum(Session.total_sales).label('total_sales')
            ).join(
                Session, Session.user_id == User.id
            ).filter(
                Session.login_time >= start_date
            ).group_by(
                User.id
            ).order_by(
                func.sum(Session.total_sales).desc()
            ).all()
            
            # Update table
            self.employee_table.setRowCount(len(results))
            
            employees = []
            sales = []
            
            for idx, row in enumerate(results):
                display_name = f"{row.full_name} ({row.username})"
                self.employee_table.setItem(idx, 0, QTableWidgetItem(display_name))
                self.employee_table.setItem(idx, 1, QTableWidgetItem(str(row.session_count)))
                
                orders = row.total_orders or 0
                total = row.total_sales or 0
                avg = total / orders if orders > 0 else 0
                
                self.employee_table.setItem(idx, 2, QTableWidgetItem(str(orders)))
                self.employee_table.setItem(idx, 3, QTableWidgetItem(CurrencyFormatter.format_lyd(total)))
                self.employee_table.setItem(idx, 4, QTableWidgetItem(CurrencyFormatter.format_lyd(avg)))
                
                employees.append(row.full_name[:10])
                sales.append(CurrencyFormatter.fils_to_lyd(total))
            
            # Update chart - clear all series and axes first
            self.employee_chart.removeAllSeries()
            for axis in self.employee_chart.axes():
                self.employee_chart.removeAxis(axis)
            
            bar_set = QBarSet("إجمالي المبيعات (د.ل)")
            for sale in sales:
                bar_set.append(sale)
            
            series = QBarSeries()
            series.append(bar_set)
            
            self.employee_chart.addSeries(series)
            
            axis_x = QBarCategoryAxis()
            axis_x.append(employees)
            self.employee_chart.addAxis(axis_x, Qt.AlignmentFlag.AlignBottom)
            series.attachAxis(axis_x)
            
            axis_y = QValueAxis()
            axis_y.setTitleText("المبيعات (د.ل)")
            self.employee_chart.addAxis(axis_y, Qt.AlignmentFlag.AlignLeft)
            series.attachAxis(axis_y)
            
        finally:
            db.close()
            
    def export_money_excel(self):
        """Export money report to Excel"""
        filename, _ = QFileDialog.getSaveFileName(
            self, "تصدير تقرير الإيرادات", "money_report.xlsx", "ملفات إكسل (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "تحليل الإيرادات"
            
            # Headers
            headers = ["التاريخ", "الطلبات", "مبيعات نقدية (د.ل)", "مبيعات بطاقات (د.ل)", "الإجمالي (د.ل)"]
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row in range(self.money_table.rowCount()):
                ws.append([
                    self.money_table.item(row, 0).text(),
                    int(self.money_table.item(row, 1).text()),
                    self.money_table.item(row, 2).text(),
                    self.money_table.item(row, 3).text(),
                    self.money_table.item(row, 4).text()
                ])
            
            wb.save(filename)
            QMessageBox.information(self, "تم بنجاح", f"تم تصدير التقرير إلى {filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"تعذر التصدير: {e}")
            
    def export_popular_excel(self):
        """Export popular items to Excel"""
        filename, _ = QFileDialog.getSaveFileName(
            self, "تصدير الأصناف الأكثر مبيعًا", "popular_items.xlsx", "ملفات إكسل (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "الأصناف الأكثر مبيعًا"
            
            headers = ["الترتيب", "الصنف", "الكمية المباعة", "الإيراد (د.ل)"]
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            for row in range(self.popular_table.rowCount()):
                ws.append([
                    self.popular_table.item(row, 0).text(),
                    self.popular_table.item(row, 1).text(),
                    int(self.popular_table.item(row, 2).text()),
                    self.popular_table.item(row, 3).text()
                ])
            
            wb.save(filename)
            QMessageBox.information(self, "تم بنجاح", f"تم تصدير التقرير إلى {filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"تعذر التصدير: {e}")
            
    def export_peak_excel(self):
        """Export peak hours to Excel"""
        filename, _ = QFileDialog.getSaveFileName(
            self, "تصدير ساعات الذروة", "peak_hours.xlsx", "ملفات إكسل (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ساعات الذروة"
            
            headers = ["الساعة", "الطلبات", "الإيراد (د.ل)"]
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            for row in range(self.peak_table.rowCount()):
                ws.append([
                    self.peak_table.item(row, 0).text(),
                    int(self.peak_table.item(row, 1).text()),
                    self.peak_table.item(row, 2).text()
                ])
            
            wb.save(filename)
            QMessageBox.information(self, "تم بنجاح", f"تم تصدير التقرير إلى {filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"تعذر التصدير: {e}")
            
    def export_employee_excel(self):
        """Export employee performance to Excel"""
        filename, _ = QFileDialog.getSaveFileName(
            self, "تصدير أداء الموظفين", "employee_performance.xlsx", "ملفات إكسل (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "أداء الموظفين"
            
            headers = ["الموظف", "الجلسات", "الطلبات", "إجمالي المبيعات (د.ل)", "متوسط الطلب (د.ل)"]
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            for row in range(self.employee_table.rowCount()):
                ws.append([
                    self.employee_table.item(row, 0).text(),
                    int(self.employee_table.item(row, 1).text()),
                    int(self.employee_table.item(row, 2).text()),
                    self.employee_table.item(row, 3).text(),
                    self.employee_table.item(row, 4).text()
                ])
            
            wb.save(filename)
            QMessageBox.information(self, "تم بنجاح", f"تم تصدير التقرير إلى {filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"تعذر التصدير: {e}")
